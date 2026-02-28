"""
機械学習を用いた ICL サイズ予測モデルの構築と評価 v2
==================================================
目的変数: size（実際に選択されたICLサイズ）
特徴量:
  - LV_SliceNo: 0 (Angle: 180-0)
  - ACD[Endo.]_CCT/ACD
  - CCT_SliceNo: 0 (Angle: 180-0)
  - ACW_SliceNo: 0 (Angle: 180-0)
評価: Leave-One-Out Cross-Validation (LOOCV)
予測方法: argmax（最高確率）+ 加重平均（確率分布）の両方を比較
"""

import numpy as np
import pandas as pd
import re
import os
import time
import pickle
import warnings
warnings.filterwarnings('ignore')

from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import RandomForestClassifier
from sklearn.svm import SVC
from sklearn.preprocessing import RobustScaler, label_binarize
from sklearn.model_selection import LeaveOneOut
from sklearn.metrics import (
    confusion_matrix,
    precision_recall_fscore_support,
    roc_auc_score
)
from sklearn.base import clone
from collections import Counter
import matplotlib.pyplot as plt
import seaborn as sns

try:
    from xgboost import XGBClassifier
    HAS_XGB = True
except ImportError:
    HAS_XGB = False

try:
    from lightgbm import LGBMClassifier
    HAS_LGBM = True
except ImportError:
    HAS_LGBM = False

try:
    from catboost import CatBoostClassifier
    HAS_CAT = True
except ImportError:
    HAS_CAT = False

# ============================================================
# 設定
# ============================================================
SEED = 2025
np.random.seed(SEED)

SAVE_DIR = '/content/drive/MyDrive/df/IPCL'

SIZE_CANDIDATES = [11.00, 11.25, 11.50, 11.75, 12.00, 12.25, 12.50, 12.75,
                   13.00, 13.25, 13.50, 13.75, 14.00]
SIZE_ARRAY = np.array(SIZE_CANDIDATES)
SIZE_TO_IDX = {s: i for i, s in enumerate(SIZE_CANDIDATES)}
IDX_TO_SIZE = {i: s for i, s in enumerate(SIZE_CANDIDATES)}

# ============================================================
# ユーティリティ関数
# ============================================================
def clean_feature_name(name):
    cleaned = re.sub(r'[^a-zA-Z0-9_]', '_', name)
    cleaned = re.sub(r'_+', '_', cleaned)
    return cleaned.strip('_')

def size_to_nearest(s):
    return min(SIZE_CANDIDATES, key=lambda x: abs(x - s))

def prob_to_weighted_size(prob_matrix):
    """確率分布の加重平均でサイズを予測し、最近傍0.25mm刻みに丸める"""
    weighted = prob_matrix @ SIZE_ARRAY          # 各症例の加重平均サイズ
    rounded = np.array([size_to_nearest(v) for v in weighted])
    return weighted, rounded

def evaluate_sizes(pred_sizes, actual_sizes):
    """評価指標を計算"""
    exact = np.mean(pred_sizes == actual_sizes) * 100
    w025  = np.mean(np.abs(pred_sizes - actual_sizes) <= 0.25) * 100
    w050  = np.mean(np.abs(pred_sizes - actual_sizes) <= 0.50) * 100
    return exact, w025, w050

# ============================================================
# データ準備
# ============================================================
selected_features = [
    'LV_SliceNo: 0 (Angle: 180-0)',
    'ACD[Endo.]_CCT/ACD',
    'CCT_SliceNo: 0 (Angle: 180-0)',
    'ACW_SliceNo: 0 (Angle: 180-0)',
]
original_names = selected_features.copy()

df = pd.read_excel(os.path.join(SAVE_DIR, 'IPCL3.xlsx'))

missing = [c for c in selected_features + ['size'] if c not in df.columns]
if missing:
    raise ValueError(f"以下の列が見つかりません: {missing}")

df_clean = df[selected_features + ['size']].dropna().copy()
n_excluded = len(df) - len(df_clean)

df_clean['size_label'] = df_clean['size'].apply(
    lambda s: SIZE_TO_IDX[size_to_nearest(s)])

X_raw = df_clean[selected_features].copy()
X_raw.columns = [clean_feature_name(col) for col in X_raw.columns]
feat_names = list(X_raw.columns)

y = df_clean['size_label'].values
actual_sizes = df_clean['size'].values
n = len(X_raw)

present_classes = sorted(np.unique(y))
N_CLASSES = len(SIZE_CANDIDATES)

print("=" * 70)
print("機械学習を用いた ICL サイズ予測モデルの構築と評価 v2")
print("=" * 70)
print(f"データファイル: IPCL3.xlsx")
print(f"サンプル数: {n}（欠損除外: {n_excluded}例）")
print(f"特徴量: {original_names}")
print(f"予測方法: argmax + 加重平均（確率分布）")
print(f"評価: LOOCV (n={n})")
print(f"\nサイズ分布:")
for s in SIZE_CANDIDATES:
    idx = SIZE_TO_IDX[s]
    cnt = np.sum(y == idx)
    bar = '█' * cnt if cnt > 0 else '─'
    print(f"  {s:.2f}mm: {cnt:2d}例  {bar}")
print("=" * 70)

# ============================================================
# Part A: LOOCVモデル評価
# ============================================================
print(f"\n{'='*70}")
print("【Part A】分類モデル ─ ICLサイズ予測（LOOCV）")
print(f"{'='*70}")

def get_classification_models():
    models = {}
    models['LogReg'] = LogisticRegression(
        C=1.0, max_iter=1000, random_state=SEED,
        solver='lbfgs', multi_class='auto')
    models['SVM'] = SVC(
        C=10.0, kernel='rbf', gamma='scale',
        probability=True, random_state=SEED)
    models['RF'] = RandomForestClassifier(
        n_estimators=200, max_depth=4, min_samples_leaf=3,
        random_state=SEED, n_jobs=-1)
    if HAS_XGB:
        models['XGBoost'] = XGBClassifier(
            n_estimators=100, max_depth=3, learning_rate=0.05,
            subsample=0.8, colsample_bytree=0.8,
            reg_alpha=0.1, reg_lambda=1.0,
            eval_metric='mlogloss',
            random_state=SEED, verbosity=0, n_jobs=-1)
    if HAS_LGBM:
        models['LightGBM'] = LGBMClassifier(
            n_estimators=100, max_depth=3, learning_rate=0.05,
            num_leaves=15, subsample=0.8, colsample_bytree=0.8,
            random_state=SEED, verbose=-1, n_jobs=-1)
    if HAS_CAT:
        models['CatBoost'] = CatBoostClassifier(
            iterations=100, depth=3, learning_rate=0.05,
            l2_leaf_reg=3.0, bootstrap_type='Bernoulli', subsample=0.8,
            random_seed=SEED, verbose=0)
    return models

clf_models = get_classification_models()
clf_results = {}
loo = LeaveOneOut()

start_all = time.time()
print(f"\n{len(clf_models)}モデルをLOOCVで評価中...\n")
print(f"  {'Model':<12s}  {'Exact(argmax)':>14} {'±0.25(argmax)':>14} "
      f"{'Exact(wavg)':>12} {'±0.25(wavg)':>12}  {'AUC':>7}")
print(f"  {'-'*80}")

for model_name, model_template in clf_models.items():
    start = time.time()
    y_pred_argmax = np.zeros(n, dtype=int)
    y_prob = np.zeros((n, N_CLASSES))

    for train_idx, test_idx in loo.split(X_raw):
        X_tr = X_raw.iloc[train_idx]
        X_te = X_raw.iloc[test_idx]
        y_tr = y[train_idx]

        sc = RobustScaler()
        X_tr_s = sc.fit_transform(X_tr)
        X_te_s = sc.transform(X_te)

        mdl = clone(model_template)
        mdl.fit(X_tr_s, y_tr)

        y_pred_argmax[test_idx[0]] = int(mdl.predict(X_te_s)[0])

        proba = mdl.predict_proba(X_te_s)[0]
        for ci, cls in enumerate(mdl.classes_):
            y_prob[test_idx[0], cls] = proba[ci]

    elapsed = time.time() - start

    # argmax評価
    pred_sizes_argmax = np.array([IDX_TO_SIZE[i] for i in y_pred_argmax])
    exact_am, w025_am, w050_am = evaluate_sizes(pred_sizes_argmax, actual_sizes)

    # 加重平均評価
    _, pred_sizes_wavg = prob_to_weighted_size(y_prob)
    exact_wv, w025_wv, w050_wv = evaluate_sizes(pred_sizes_wavg, actual_sizes)

    # AUC
    y_bin = label_binarize(y, classes=present_classes)
    auc_scores = []
    for i, cls in enumerate(present_classes):
        if len(np.unique(y_bin[:, i])) > 1:
            try:
                auc_scores.append(roc_auc_score(y_bin[:, i], y_prob[:, cls]))
            except:
                auc_scores.append(np.nan)
        else:
            auc_scores.append(np.nan)
    valid_aucs = [a for a in auc_scores if not np.isnan(a)]
    macro_auc = np.mean(valid_aucs) if valid_aucs else np.nan

    clf_results[model_name] = {
        # argmax
        'y_pred_argmax': y_pred_argmax,
        'pred_sizes_argmax': pred_sizes_argmax,
        'exact_argmax': exact_am,
        'within_025_argmax': w025_am,
        'within_050_argmax': w050_am,
        # 加重平均
        'pred_sizes_wavg': pred_sizes_wavg,
        'exact_wavg': exact_wv,
        'within_025_wavg': w025_wv,
        'within_050_wavg': w050_wv,
        # 共通
        'y_prob': y_prob,
        'macro_auc': macro_auc,
        # 後方互換（既存の外部検証コード用）
        'exact': exact_am,
        'within_025': w025_am,
        'within_050': w050_am,
        'time': elapsed,
    }

    print(f"  {model_name:<12s}  {exact_am:>12.1f}%  {w025_am:>12.1f}%  "
          f"{exact_wv:>10.1f}%  {w025_wv:>10.1f}%  {macro_auc:>7.3f}  ({elapsed:.1f}s)")

total_clf = time.time() - start_all
print(f"\n分類モデル完了: {total_clf:.1f}秒")

# ============================================================
# アンサンブル（トップ2: ±0.25mm_argmax基準）
# ============================================================
sorted_clf = sorted(clf_results.items(),
                    key=lambda x: x[1]['within_025_argmax'], reverse=True)
top2_names = [sorted_clf[0][0], sorted_clf[1][0]]
print(f"\nトップ2モデル: {top2_names[0]} + {top2_names[1]} でアンサンブル構築中...")

# 確率の平均アンサンブル
ens_prob = (clf_results[top2_names[0]]['y_prob'] +
            clf_results[top2_names[1]]['y_prob']) / 2

# argmax
ens_pred_argmax = np.argmax(ens_prob, axis=1)
ens_sizes_argmax = np.array([IDX_TO_SIZE[i] for i in ens_pred_argmax])
ens_exact_am, ens_025_am, ens_050_am = evaluate_sizes(ens_sizes_argmax, actual_sizes)

# 加重平均
_, ens_sizes_wavg = prob_to_weighted_size(ens_prob)
ens_exact_wv, ens_025_wv, ens_050_wv = evaluate_sizes(ens_sizes_wavg, actual_sizes)

# AUC
y_bin = label_binarize(y, classes=present_classes)
auc_ens = []
for i, cls in enumerate(present_classes):
    if len(np.unique(y_bin[:, i])) > 1:
        try:
            auc_ens.append(roc_auc_score(y_bin[:, i], ens_prob[:, cls]))
        except:
            auc_ens.append(np.nan)
    else:
        auc_ens.append(np.nan)
macro_auc_ens = np.mean([a for a in auc_ens if not np.isnan(a)])

ensemble_name = f'Ensemble ({top2_names[0]}+{top2_names[1]})'
clf_results[ensemble_name] = {
    'y_pred_argmax': ens_pred_argmax,
    'pred_sizes_argmax': ens_sizes_argmax,
    'exact_argmax': ens_exact_am,
    'within_025_argmax': ens_025_am,
    'within_050_argmax': ens_050_am,
    'pred_sizes_wavg': ens_sizes_wavg,
    'exact_wavg': ens_exact_wv,
    'within_025_wavg': ens_025_wv,
    'within_050_wavg': ens_050_wv,
    'y_prob': ens_prob,
    'macro_auc': macro_auc_ens,
    'exact': ens_exact_am,
    'within_025': ens_025_am,
    'within_050': ens_050_am,
    'time': 0,
}

print(f"  {ensemble_name:<30s}  Exact(argmax)={ens_exact_am:.1f}%  "
      f"±0.25(argmax)={ens_025_am:.1f}%  "
      f"Exact(wavg)={ens_exact_wv:.1f}%  ±0.25(wavg)={ens_025_wv:.1f}%")

# 最良モデル（±0.25mm_argmax基準）
sorted_clf = sorted(clf_results.items(),
                    key=lambda x: x[1]['within_025_argmax'], reverse=True)
best_name = sorted_clf[0][0]
best_res  = sorted_clf[0][1]

print(f"\n★ 最良モデル: {best_name}")
print(f"  argmax: Exact={best_res['exact_argmax']:.1f}%  "
      f"±0.25mm={best_res['within_025_argmax']:.1f}%  "
      f"±0.50mm={best_res['within_050_argmax']:.1f}%")
print(f"  wavg  : Exact={best_res['exact_wavg']:.1f}%  "
      f"±0.25mm={best_res['within_025_wavg']:.1f}%  "
      f"±0.50mm={best_res['within_050_wavg']:.1f}%")
print(f"  AUC   : {best_res['macro_auc']:.3f}")

# ============================================================
# Part B: 最終モデル構築（全データ）
# ============================================================
print(f"\n{'='*70}")
print("【Part B】最終モデル構築（全データ）")
print(f"{'='*70}")

scaler_final = RobustScaler()
X_final = scaler_final.fit_transform(X_raw)

final_models = {}
for name in top2_names:
    mdl = clone(clf_models[name])
    mdl.fit(X_final, y)
    final_models[name] = mdl
    print(f"  {name}: fit完了")

# ============================================================
# Part C: 比較表
# ============================================================
print(f"\n{'='*70}")
print("【Part C】総合比較表（±0.25mm_argmax順）")
print(f"{'='*70}")
print(f"\n  {'Model':<30s} {'Exact':>7} {'±0.25':>7} {'±0.50':>7}  "
      f"{'Exact(w)':>9} {'±0.25(w)':>9} {'±0.50(w)':>9}  {'AUC':>7}")
print(f"  {'-'*90}")
for name, r in sorted_clf:
    print(f"  {name:<30s} "
          f"{r['exact_argmax']:>6.1f}%  {r['within_025_argmax']:>6.1f}%  {r['within_050_argmax']:>6.1f}%  "
          f"{r['exact_wavg']:>8.1f}%  {r['within_025_wavg']:>8.1f}%  {r['within_050_wavg']:>8.1f}%  "
          f"{r['macro_auc']:>7.3f}")

# ============================================================
# Part D: 可視化
# ============================================================
print(f"\n図表作成中...")

best_pred_am   = best_res['y_pred_argmax']
best_sizes_am  = best_res['pred_sizes_argmax']
best_sizes_wv  = best_res['pred_sizes_wavg']
best_prob      = best_res['y_prob']

fig, axes = plt.subplots(2, 2, figsize=(16, 12))

# (a) Confusion Matrix - argmax
ax = axes[0, 0]
cm_am = confusion_matrix(y, best_pred_am, labels=present_classes)
size_labels = [f"{IDX_TO_SIZE[c]:.2f}" for c in present_classes]
sns.heatmap(cm_am, annot=True, fmt='d', cmap='Blues',
            xticklabels=size_labels, yticklabels=size_labels,
            ax=ax, linewidths=0.5, linecolor='gray', annot_kws={'size': 9})
ax.set_xlabel('Predicted Size (mm)', fontweight='bold')
ax.set_ylabel('Actual Size (mm)', fontweight='bold')
ax.set_title(f'(a) Confusion Matrix: argmax\n{best_name} | '
             f'Exact={best_res["exact_argmax"]:.1f}%, ±0.25mm={best_res["within_025_argmax"]:.1f}%',
             fontweight='bold')

# (b) Confusion Matrix - 加重平均
ax = axes[0, 1]
best_pred_wv_idx = np.array([SIZE_TO_IDX[s] for s in best_sizes_wv])
cm_wv = confusion_matrix(y, best_pred_wv_idx, labels=present_classes)
sns.heatmap(cm_wv, annot=True, fmt='d', cmap='Oranges',
            xticklabels=size_labels, yticklabels=size_labels,
            ax=ax, linewidths=0.5, linecolor='gray', annot_kws={'size': 9})
ax.set_xlabel('Predicted Size (mm)', fontweight='bold')
ax.set_ylabel('Actual Size (mm)', fontweight='bold')
ax.set_title(f'(b) Confusion Matrix: Weighted Average\n{best_name} | '
             f'Exact={best_res["exact_wavg"]:.1f}%, ±0.25mm={best_res["within_025_wavg"]:.1f}%',
             fontweight='bold')

# (c) 確率分布ヒートマップ（全症例）
ax = axes[1, 0]
# 実際のサイズ順にソート
sort_idx = np.argsort(actual_sizes)
prob_sorted = best_prob[sort_idx][:, present_classes]
size_cols = [f"{IDX_TO_SIZE[c]:.2f}" for c in present_classes]
im = ax.imshow(prob_sorted.T, aspect='auto', cmap='YlOrRd', vmin=0, vmax=1)
ax.set_xlabel('Case (sorted by actual size)', fontweight='bold')
ax.set_ylabel('Predicted Size (mm)', fontweight='bold')
ax.set_yticks(range(len(present_classes)))
ax.set_yticklabels(size_cols, fontsize=8)
plt.colorbar(im, ax=ax, label='Probability')
ax.set_title(f'(c) Probability Distribution Heatmap\n{best_name}', fontweight='bold')

# (d) argmax vs 加重平均の精度比較
ax = axes[1, 1]
categories = ['Exact', '±0.25mm', '±0.50mm']
am_vals = [best_res['exact_argmax'], best_res['within_025_argmax'], best_res['within_050_argmax']]
wv_vals = [best_res['exact_wavg'],   best_res['within_025_wavg'],   best_res['within_050_wavg']]
x_pos = np.arange(3)
bars1 = ax.bar(x_pos - 0.2, am_vals, 0.38, label='argmax',
               color='#2196F3', alpha=0.85, edgecolor='black', linewidth=0.5)
bars2 = ax.bar(x_pos + 0.2, wv_vals, 0.38, label='Weighted Average',
               color='#FF9800', alpha=0.85, edgecolor='black', linewidth=0.5)
ax.set_xticks(x_pos)
ax.set_xticklabels(categories, fontweight='bold')
ax.set_ylabel('Accuracy (%)', fontweight='bold')
ax.set_title(f'(d) argmax vs Weighted Average\n{best_name} (LOOCV)', fontweight='bold')
ax.legend(fontsize=10); ax.grid(True, alpha=0.2, axis='y')
ax.set_ylim(0, 115)
for bar, v in zip(list(bars1) + list(bars2), am_vals + wv_vals):
    ax.text(bar.get_x() + bar.get_width()/2, v + 1,
            f'{v:.1f}%', ha='center', fontsize=9, fontweight='bold')

fig.suptitle(f'ICL Size Prediction: argmax vs Weighted Average (LOOCV, n={n})',
             fontsize=13, fontweight='bold')
fig.tight_layout()
fig.savefig(os.path.join(SAVE_DIR, 'Fig_size_model_comparison_v2.png'),
            dpi=300, bbox_inches='tight')
plt.show()

# ============================================================
# pickle保存
# ============================================================
metrics_best = {
    'accuracy':     best_res['exact_argmax'],
    'within_025':   best_res['within_025_argmax'],
    'within_050':   best_res['within_050_argmax'],
    'exact_wavg':   best_res['exact_wavg'],
    'within_025_wavg': best_res['within_025_wavg'],
    'within_050_wavg': best_res['within_050_wavg'],
    'macro_auc':    best_res['macro_auc'],
}

save_data = {
    'final_models': final_models,
    'scaler': scaler_final,
    'best_model_name': best_name,
    'top2_names': top2_names,
    'feature_names_original': original_names,
    'feature_names_clean': feat_names,
    'size_candidates': SIZE_CANDIDATES,
    'size_to_idx': SIZE_TO_IDX,
    'idx_to_size': IDX_TO_SIZE,
    'present_classes': present_classes,
    'loocv_results': clf_results,
    'metrics': metrics_best,
    'X_train_scaled': X_final,
    'y_train': y,
    'actual_sizes_train': actual_sizes,
    'n_train': n,
}

pkl_path = os.path.join(SAVE_DIR, 'vault_size_prediction.pkl')
try:
    with open(pkl_path, 'wb') as f:
        pickle.dump(save_data, f)
    print(f"\nモデルを {pkl_path} に保存しました。")
except:
    with open('vault_size_prediction.pkl', 'wb') as f:
        pickle.dump(save_data, f)
    print(f"\nモデルをローカルに保存しました。")

# ============================================================
# Excel出力
# ============================================================
print("\nTable (Excel) 作成中...")
try:
    xlsx_path = os.path.join(SAVE_DIR, 'Table_size_results_v2.xlsx')
    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:

        # モデル比較
        model_table = []
        for name, r in sorted_clf:
            model_table.append({
                'Model': name,
                'Exact_argmax (%)':   round(r['exact_argmax'], 1),
                '±0.25mm_argmax (%)': round(r['within_025_argmax'], 1),
                '±0.50mm_argmax (%)': round(r['within_050_argmax'], 1),
                'Exact_wavg (%)':     round(r['exact_wavg'], 1),
                '±0.25mm_wavg (%)':   round(r['within_025_wavg'], 1),
                '±0.50mm_wavg (%)':   round(r['within_050_wavg'], 1),
                'Macro AUC':          round(r['macro_auc'], 3),
            })
        pd.DataFrame(model_table).to_excel(writer, 'Model Comparison', index=False)

        # 最良モデル予測結果（argmax + wavg）
        pred_table = pd.DataFrame({
            'Case':              np.arange(1, n + 1),
            'Actual_Size':       actual_sizes,
            'Pred_argmax':       best_sizes_am,
            'Diff_argmax':       best_sizes_am - actual_sizes,
            'Exact_argmax':      best_sizes_am == actual_sizes,
            'Within025_argmax':  np.abs(best_sizes_am - actual_sizes) <= 0.25,
            'Pred_wavg':         best_sizes_wv,
            'Diff_wavg':         best_sizes_wv - actual_sizes,
            'Exact_wavg':        best_sizes_wv == actual_sizes,
            'Within025_wavg':    np.abs(best_sizes_wv - actual_sizes) <= 0.25,
        })
        pred_table.to_excel(writer, 'LOOCV Predictions', index=False)

        # 確率分布
        prob_df = pd.DataFrame(
            best_prob[:, present_classes],
            columns=[f"{IDX_TO_SIZE[c]:.2f}mm" for c in present_classes]
        )
        prob_df.insert(0, 'Actual_Size', actual_sizes)
        prob_df.insert(1, 'Pred_argmax', best_sizes_am)
        prob_df.insert(2, 'Pred_wavg',   best_sizes_wv)
        prob_df.to_excel(writer, 'Probability Distribution', index=False)

    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    wb = load_workbook(xlsx_path)
    hdr_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    hdr_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    cell_font = Font(name='Arial', size=10)
    thin = Border(*(Side(style='thin'),)*4)
    center = Alignment(horizontal='center', vertical='center')
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = hdr_font; cell.fill = hdr_fill
            cell.alignment = center; cell.border = thin
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = cell_font; cell.alignment = center; cell.border = thin
        for col in ws.columns:
            ml = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max(ml + 3, 12)
    wb.save(xlsx_path)
    print(f"  → {xlsx_path}")
except Exception as e:
    print(f"  Excel error: {e}")

# ============================================================
# 最終サマリー
# ============================================================
print(f"\n{'='*70}")
print("【最終サマリー】")
print(f"{'='*70}")
print(f"\n■ 最良モデル: {best_name}")
print(f"\n  {'指標':<15s} {'argmax':>10} {'加重平均':>10}")
print(f"  {'-'*38}")
print(f"  {'Exact (%)':<15s} {best_res['exact_argmax']:>10.1f} {best_res['exact_wavg']:>10.1f}")
print(f"  {'±0.25mm (%)':<15s} {best_res['within_025_argmax']:>10.1f} {best_res['within_025_wavg']:>10.1f}")
print(f"  {'±0.50mm (%)':<15s} {best_res['within_050_argmax']:>10.1f} {best_res['within_050_wavg']:>10.1f}")
print(f"  {'Macro AUC':<15s} {best_res['macro_auc']:>10.3f} {'―':>10}")
print(f"\n■ 出力ファイル:")
print(f"  Fig_size_model_comparison_v2.png")
print(f"  Table_size_results_v2.xlsx")
print(f"  vault_size_prediction.pkl")
print(f"\n■ 計算時間: {time.time() - start_all:.1f}秒")
print(f"\n{'='*70}")
print("分析完了")
print(f"{'='*70}")
